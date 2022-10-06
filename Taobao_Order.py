import json
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
from openpyxl.styles import Side, Font, Alignment, Border
from decimal import Decimal


class Taobao(object):
    def __init__(self):
        options = webdriver.ChromeOptions()
        options.add_argument('--headless')
        self.driver = webdriver.Chrome(options=options)
        self.driver.get('https://buyertrade.taobao.com/trade/itemlist/list_bought_items.htm?')
        self.driver.maximize_window()   # 放大窗口

        self.date =[]        # 存放日期
        self.order_good = []        # 存放订单号
        self.good_name = []        # 存放商品名
        self.num =[]        # 存放购买数量
        self.prices = []        # 存放商品价格
        self.stores = []        # 存放店铺名

    # 登录
    def login(self):
        # 获取保存下的cookie值
        with open('taobao_cookies.txt', 'r', encoding='utf8') as f:
            listCookies = json.loads(f.read())

        # 往browser里添加cookies
        for cookie in listCookies:
            cookie_dict = {
                'domain': '.taobao.com',
                'name': cookie.get('name'),
                'value': cookie.get('value'),
                "expires": '',
                'path': '/',
                'httpOnly': False,
                'HostOnly': False,
                'Secure': False
            }
            self.driver.add_cookie(cookie_dict)

        self.driver.refresh()

        # 等待快速登录按钮出现并点击
        WebDriverWait(self.driver, 1000).until(
            EC.presence_of_element_located((By.XPATH, '//div[@class="fm-btn"]/button'))
        )
        self.driver.find_element(By.XPATH, '//div[@class="fm-btn"]/button').click()

        print('登录成功！请等待获取信息……')

    # 解析数据
    def parse_data(self):
        # 点击交易状态元素出现并点击
        WebDriverWait(self.driver, 1000).until(
            EC.presence_of_element_located((By.XPATH, '//span[@class="rc-select-selection__rendered"]/span[text()="交易状态"]'))
        )
        self.driver.find_element(By.XPATH, '//span[@class="rc-select-selection__rendered"]/span[text()="交易状态"]').click()

        # 点击交易成功
        self.driver.find_element(By.XPATH, '//li[text()="交易成功"]').click()
        time.sleep(3)

        page = 0
        while True:
            try:
                # 等待日期元素出现
                WebDriverWait(self.driver, 1000).until(
                    EC.presence_of_element_located(
                        (By.XPATH, '//*[@id="tp-bought-root"]/div[*]/div/table/tbody[1]/tr/td[2]/span/a'))
                )
                time.sleep(2)   # 时间等待
                # 获取日期
                self.date += [i.text for i in self.driver.find_elements(By.XPATH, '//*[@id="tp-bought-root"]/div[*]/div/table/tbody[1]/tr/td[1]/label/span[2]')]
                # 获取订单号 text
                self.order_good += [i.text for i in self.driver.find_elements(By.XPATH, '//*[@id="tp-bought-root"]/div[*]/div/table/tbody[1]/tr/td[1]/span/span[3]')]
                # 获取商品名 text
                self.good_name += [i.text for i in self.driver.find_elements(By.XPATH, '//*[@id="tp-bought-root"]/div[*]/div/table/tbody[2]/tr[1]/td[1]/div/div[2]/p[1]/a[1]/span[2]')]
                # 获取购买数量 text
                self.num += [int(i.text) for i in self.driver.find_elements(By.XPATH, '//*[@id="tp-bought-root"]/div[*]/div/table/tbody[2]/tr[1]/td[3]/div/p')]
                # 获取商品价格，保留两位小数
                self.prices += [Decimal(str(round(float(i.text), 2))) for i in self.driver.find_elements(By.XPATH, '//*[@id="tp-bought-root"]/div[*]/div/table/tbody[2]/tr[1]/td[5]/div/div[1]/p/strong/span[2]')]
                # 获取店铺名
                self.stores += [i.text for i in self.driver.find_elements(By.XPATH, '//*[@id="tp-bought-root"]/div[*]/div/table/tbody[1]/tr/td[2]/span/a')]

                page += 1
                print(f'第{page}页订单信息获取成功')

                # 判断下一页是否可以点击
                next_page = self.driver.find_element(By.XPATH, '//button[text()="下一页"]')
                if next_page.is_enabled(): # 如果可以点击，则点击
                    next_page.click()
                else:   # 不可点击则退出循环
                    break
            except:
                break

    # 将订单信息写入文件
    def save_file(self):
        f = openpyxl.Workbook()
        # 创建工作表
        sheet = f.create_sheet('淘宝订单')
        # 构建表头
        sheet.append(['日期', '订单号', '商品名', '购买数量', '商品价格', '店铺名'])

        # 设置字体的样式
        algn = Alignment(vertical='center', horizontal='center')
        side = Side(style='thin', color='000000')
        border = Border(right=side, left=side, top=side, bottom=side)
        font = Font(name='宋体', size=12)

        # 设置列宽
        sheet.column_dimensions['A'].width = 12
        sheet.column_dimensions['B'].width = 23
        sheet.column_dimensions['C'].width = 53
        sheet.column_dimensions['D'].width = 12
        sheet.column_dimensions['E'].width = 12
        sheet.column_dimensions['F'].width = 26

        # 获取的数据量
        num = len(self.date)

        for i in sheet[f'A1:F{num+5}']:
            for j in i:
                j.alignment = algn
                j.border = border
                j.font = font

        # 将内容写入表格
        for i in range(num):
            sheet.cell(i+2, 1).value = self.date[i]
            sheet.cell(i+2, 2).value = self.order_good[i]
            sheet.cell(i+2, 3).value = self.good_name[i]
            sheet.cell(i+2, 4).value = self.num[i]
            sheet.cell(i+2, 5).value = self.prices[i]
            sheet.cell(i+2, 6).value = self.stores[i]

        sheet.cell(num+3, 1).value = f'总购买商品样数：{num}样；总购买数量：{sum(self.num)}件；总共花费：{sum(self.prices)}元'
        # 合并单元格
        sheet.merge_cells(f'A{num+3}:C{num+5}')

        # 保存表格
        f.save('淘宝订单信息.xlsx')
        print('淘宝订单信息.xlsx 文件保存成功！')

    def run(self):
        self.login()
        self.parse_data()
        self.save_file()

taobao = Taobao()
taobao.run()


